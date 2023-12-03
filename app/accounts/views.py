import openpyxl
from django.utils.decorators import method_decorator
from rest_framework import status, generics, filters
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .constants import PENDING_COMPLETE_DATA, COMPLETE, FIRST_CLOSET
from .functions import is_active
from .models import User, Item
from .serializers import UserSerializer, CompleteProfileSerializer, ItemSerializer, ExcelFileSerializer


class UsersListAPI(APIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    @method_decorator(is_active)
    def get(self, request):
        user = request.user
        obj = User.objects.all()
        serializer = self.serializer_class(obj, many=True)
        if user.is_superuser:
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(status=status.HTTP_403_FORBIDDEN)


class UserDetailAPI(APIView):
    permission_classes = [IsAuthenticated]
    serializer_class = UserSerializer

    @staticmethod
    def get_object(pk):
        try:
            return User.objects.get(pk=pk)
        except User.DoesNotExist:
            return None

    @method_decorator(is_active)
    def get(self, request, pk):
        obj = self.get_object(pk)
        if obj is None:
            return Response(status=status.HTTP_404_NOT_FOUND)
        serializer = self.serializer_class(obj)
        if request.user.is_superuser:
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(status=status.HTTP_403_FORBIDDEN)

    @method_decorator(is_active)
    def put(self, request, pk):
        obj = self.get_object(pk)
        if obj is None:
            return Response(status=status.HTTP_404_NOT_FOUND)
        serializer = self.serializer_class(obj, data=request.data)
        if obj.id == request.user.id or request.user.is_superuser:
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response(status=status.HTTP_403_FORBIDDEN)

    @method_decorator(is_active)
    def delete(self, request, pk):
        obj = self.get_object(pk)
        if obj is None:
            return Response(status=status.HTTP_404_NOT_FOUND)
        if obj.id == request.user.id or request.user.is_superuser:
            obj.is_active = False
            obj.save()
            return Response(status=status.HTTP_200_OK)
        else:
            return Response(status=status.HTTP_403_FORBIDDEN)


class CompleteProfileAPI(APIView):
    permission_classes = [IsAuthenticated]
    serializer_class = CompleteProfileSerializer

    def put(self, request):
        user = request.user
        if user.status == PENDING_COMPLETE_DATA:
            serializer = self.serializer_class(data=request.data)
            if serializer.is_valid():
                user.first_name = serializer.validated_data['first_name']
                user.last_name = serializer.validated_data['last_name']
                user.telephone = serializer.validated_data['telephone']
                user.status = COMPLETE
                user.save()
                return Response({'user_status': COMPLETE}, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({"The User: {}, has already completed his profile"
                            .format(user)},
                            status=status.HTTP_400_BAD_REQUEST)


class ItemsListAPI(generics.ListAPIView):
    """
    List of all balls drawn.
    """
    serializer_class = ItemSerializer
    queryset = Item.objects.all()
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'closet_number']


class ItemSubmitAPI(APIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ItemSerializer

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Item submitted successfully'},
                            status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class SubmitExcel(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    serializer_class = ExcelFileSerializer

    def post(self, request):
        file_serializer = self.serializer_class(data=request.data)
        if file_serializer.is_valid() and request.FILES['file'].name.endswith('.xlsx'):
            excel_file = request.FILES['file']
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            items_data = []
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                values_only=True
            ):
                if row[0] == 'item_id':
                    continue
                items_data.append({
                    'item_id': row[0],
                    'name': row[1],
                    'description': row[2],
                    'quantity': row[3],
                    'closet_number': row[4] if row[4] in [1, 2] else FIRST_CLOSET,
                })

            old_items = Item.objects.all()
            if old_items:
                old_items.delete()

            for item_data in items_data:
                item_serializer = ItemSerializer(data=item_data)
                if item_serializer.is_valid():

                    item_serializer.save()
                else:
                    return Response(item_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            return Response({'message': 'Items submitted successfully'},
                            status=status.HTTP_200_OK)

        else:
            return Response({'message': 'File must be an excel file'},
                            status=status.HTTP_400_BAD_REQUEST)

